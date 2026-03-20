from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# Create workbook
wb = Workbook()

# Product sheet
ws = wb.active
ws.title = "products"

headers = [
    "name",
    "initial_stock",
    "cost",
    "price",
    "tracks_stock",
    "low_stock_threshold"
]

ws.append(headers)

# Style headers
for cell in ws[1]:
    cell.font = Font(bold=True)

# Column widths
widths = [30, 15, 12, 12, 15, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w

# Freeze header
ws.freeze_panes = "A2"

# Tracks_stock dropdown
dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("E2:E1000")

# Conditional formatting colors
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Duplicate names
ws.conditional_formatting.add(
    "A2:A1000",
    FormulaRule(
        formula=['COUNTIF($A:$A,A2)>1'],
        fill=red_fill
    )
)

# Missing name
ws.conditional_formatting.add(
    "A2:A1000",
    FormulaRule(
        formula=['A2=""'],
        fill=red_fill
    )
)

# Price lower than cost
ws.conditional_formatting.add(
    "D2:D1000",
    FormulaRule(
        formula=['D2<C2'],
        fill=yellow_fill
    )
)

# Instructions sheet
instructions = wb.create_sheet("instructions")

instructions["A1"] = "VENDR Product Import Template"
instructions["A1"].font = Font(bold=True)

instructions["A3"] = "Instructions:/Instrucciones"
instructions["A4"] = "1. Paste your product data into the 'products' sheet./ Pega los datos de tus productos a la hoja 'productos'."
instructions["A5"] = "2. Do NOT modify the column names./ NO MODIFICAR los nombres de columnas"
instructions["A6"] = "3. Fix any highlighted errors./ CORRIGE errores en las celdas resaltadas"
instructions["A7"] = "4. Save the file and upload it into VENDR./ Guarda el archivo y sube a VENDR"

instructions.column_dimensions["A"].width = 160

# Save template
wb.save("vendr_import_template.xlsx")

print("Template created successfully.")